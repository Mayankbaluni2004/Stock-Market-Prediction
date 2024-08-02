import openpyxl
import numpy as np
import matplotlib.pyplot as plt

def extract_values_from_sheet(sheet, k, column):
    values_range = sheet[f'{column}2:{column}{k}']
    return [cell.value if cell.value is not None else 0 for row in values_range for cell in row]

def calculate_correlation(sheet1, sheet2, k, column):
    values_list1 = extract_values_from_sheet(sheet1, k, column)
    values_list2 = extract_values_from_sheet(sheet2, k, column)
    correlation_coefficient = np.corrcoef(values_list1, values_list2)[0, 1]
    return correlation_coefficient

def process_correlation(columns):
    wb1 = openpyxl.load_workbook("Reliance Dataset.xlsx")
    wb2 = openpyxl.load_workbook("Adani Dataset.xlsx")

    sheet_names = [f'week {i}' for i in range(1, 14)]
    k = 6
    all_weekly_coefficients = []

    for sheet_name in sheet_names:
        if all(sheet_name in wb.sheetnames for wb in [wb1, wb2]):
            weekly_coefficients = []
            for col in columns:
                sheet1 = wb1[sheet_name]
                sheet2 = wb2[sheet_name]

                correlation_coefficient = calculate_correlation(sheet1, sheet2, k, col)
                weekly_coefficients.append(correlation_coefficient)

            all_weekly_coefficients.append(weekly_coefficients)
            k += 5
        else:
            print(f"Sheet {sheet_name} not found in one or both workbooks.")

    return all_weekly_coefficients

# Process different columns
columns_to_process = ['B', 'C', 'D', 'E']

weekly_coefficients = process_correlation(columns_to_process)

for week, coefficients in enumerate(weekly_coefficients, start=1):
    print(f"Week {week} Correlation Coefficients: {coefficients}")

# Plotting
week_numbers = list(range(1, 14))

column_labels = {'B': 'Open', 'C': 'Max', 'D': 'Min', 'E': 'Closing'}

for col in columns_to_process:
    correlation_values = [coefficients[columns_to_process.index(col)] for coefficients in weekly_coefficients]
    plt.plot(week_numbers, correlation_values, label=column_labels[col])

plt.title('Correlation Coefficients Over Weeks')
plt.xlabel('Week')
plt.ylabel('Correlation Coefficient')
plt.ylim(-1, 1)  # Set y-axis limits to vary between -1 and 1
plt.xticks(week_numbers)  # Set x-axis ticks to represent weeks 1 through 13
plt.legend()
plt.grid(True)
plt.show()
